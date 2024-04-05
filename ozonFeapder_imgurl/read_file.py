# encoding='utf-8

# @Time: 2024-03-30
# @File: %
#!/usr/bin/env
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter


class ReadFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.df = pd.read_excel(self.file_path)
        self.wb = Workbook()
        self.ws = self.wb.active
        for r in dataframe_to_rows(self.df, index=False, header=True):
            self.ws.append(r)
        self.add_column()

    def add_column(self):
        self.ws.insert_cols(1)
        self.ws.cell(row=1, column=1).value = '图片连接'
        self.ws.insert_cols(1)
        self.ws.cell(row=1, column=1).value = '图片'
        # self.ws.insert_cols(-1) # 插入在最后一列
        # self.ws.cell(row=1, column=self.ws.max_column).value = '平均送达时间'
        # self.ws.insert_cols(-1) # 插入在最后一列
        # self.ws.cell(row=1, column=self.ws.max_column).value = '平均送达时间'
        # clos_names = ['avgDeliveryDays', 'avgGmv', 'avgGmvOnAccDays', 'avgOrdersOnAccDays', 'avgPricel','createDate','daysInStock', 'fboStock', 'gmvSum', 'minSellerPrice', 'salesSchema', 'sellerId', 'soldCount', 'soldSum']
        # max_col = self.ws.max_column
        # for i in range(1, len(clos_names)+1):
        #     self.ws.insert_cols(max_col + i)
        #     self.ws.cell(row=1, column=max_col +
        #                  i).value = clos_names[i-1]


# if __name__ == '__main__':
#     file_path = './ozonFeapder_imgurl/data.xlsx'
#     save_path = './ozonFeapder_imgurl/data_img.xlsx'
#     cf = ReadFile(file_path)
#     cf.wb.save(save_path)
