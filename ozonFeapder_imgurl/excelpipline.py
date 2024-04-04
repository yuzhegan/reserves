# encoding='utf-8

# @Time: 2024-04-03
# @File: %
#!/usr/bin/env
from icecream import ic
import os
from openpyxl import Workbook, load_workbook


from feapder.pipelines import BasePipeline
from typing import Dict, List, Tuple
import pandas as pd


class ExcelPipeline(BasePipeline):
    """
    pipeline 是单线程的，批量保存数据的操作，不建议在这里写网络请求代码，如下载图片等
    """
    wb = Workbook()
    ws = wb.active
    df = pd.read_excel('./ozonFeapder_imgurl/data.xlsx')
    ws.append(list(df.columns))
    ws.insert_cols(1)
    ws.cell(1, 1).value = '图片链接'

    def save_items(self, table, items: List[Dict]) -> bool:
        """
        保存数据
        Args:
            table: 表名
            items: 数据，[{},{},...]

        Returns: 是否保存成功 True / False
                 若False，不会将本批数据入到去重库，以便再次入库

        """
        for item in items:
            self.ws.append(list(item.values()))
        self.ws
        print("自定义pipeline， 保存数据 >>>>")
        self.save_file()


        return True
    def save_file(self):
        self.wb.save('./ozonFeapder_imgurl/data_with_imgurl.xlsx')
        print('保存成功')

    def update_items(self, table, items: List[Dict], update_keys=Tuple) -> bool:
        """
        更新数据, 与UpdateItem配合使用，若爬虫中没使用UpdateItem，则可不实现此接口
        Args:
            table: 表名
            items: 数据，[{},{},...]
            update_keys: 更新的字段, 如 ("title", "publish_time")

        Returns: 是否更新成功 True / False
                 若False，不会将本批数据入到去重库，以便再次入库

        """
        self.wb.save('./ozonFeapder_imgurl/data_with_imgurl.xlsx')
        print('保存成功')

        print("自定义pipeline， 保存数据 >>>>", table, items, update_keys)

        return True

