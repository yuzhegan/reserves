# encoding='utf-8
# @Time: 2023-12-30
# @File: %
#!/usr/bin/env

import asyncio
import multiprocessing
from curl_cffi import requests
from curl_cffi.requests import AsyncSession
from lxml import etree
import logging
from parseData import *
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO
import os
from read_file import AsyncReadFile

from cookie_pool import GenCookie

class Ozon_Spider:
    def __init__(self, url):
        self.url = url
        self.headers = {
            # ... (same as before)
        }

    async def searchResultsV2(self, cookies, ua, headers, session):
        print("searchResultsV2")
        try:
            response = await session.get(self.url, cookies=cookies, headers=headers, impersonate=ua, timeout=30)
            html = etree.HTML(response.text)
            data = html.xpath('//link[@rel="preload"]/@href')
            if not data:
                data = html.xpath('//div[@class="rj"]/img/@src')
                if not data:
                    data = html.xpath('//div[@class="jn6 j6n"]//img/@src')
            return data[0] if data else None
        except Exception as e:
            logging.error(f'Error getting image URL: {e}')
            return None

async def download_image(url, headers, index, session, image_data):
    print("download_image")
    try:
        if url:
            img_url = url.replace('wc50', 'wc1000').replace('wc10000', 'wc1000')
            logging.info(f'Downloading image {index-1}: {img_url}')
            response = await session.get(img_url, headers=headers, timeout=30)
            img = Image.open(BytesIO(response.content))
            img = img.resize((150, 150), Image.LANCZOS)
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format='PNG')
            image_data[index] = (BytesIO(img_byte_arr.getvalue()), img_url)
    except Exception as e:
        logging.error(f'Error downloading image {index-1}: {e}')

async def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

    gencookies = GenCookie()
    logging.info('Start getting cookies')
    cookies, ua, headers = await gencookies.gen_cookie()
    logging.info('Finish getting cookies')

    file_path = r"./ozon_async_process/data.xlsx"
    async_read_file = AsyncReadFile(file_path)
    df, wb, ws = await async_read_file.read_file()

    manager = multiprocessing.Manager()
    image_data = manager.dict()  # 创建共享字典

    async with AsyncSession() as session:
        tasks = []
        for index, row in enumerate(df.itertuples(), start=2):
            await asyncio.sleep(1)
            url = getattr(row, '商品链接')
            ozon_spider = Ozon_Spider(url)
            task = asyncio.create_task(ozon_spider.searchResultsV2(cookies, ua, headers, session))
            tasks.append(task)

        image_urls = await asyncio.gather(*tasks)

        img_tasks = []
        for index, url in enumerate(image_urls, start=2):
            task = asyncio.create_task(download_image(url, headers, index, session, image_data))
            img_tasks.append(task)

        await asyncio.gather(*img_tasks)

    # 从共享字典中取出图片数据并写入工作表
    for index, data in image_data.items():
        img_bytes, img_url = data
        openpyxl_img = OpenpyxlImage(img_bytes)
        cell = 'A' + str(index)
        ws.add_image(openpyxl_img, cell)
        ws.cell(row=index, column=2).value = img_url

    # Set column width and row height
    ws.column_dimensions[get_column_letter(1)].width = 20
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 150 * 0.75

    # Save the workbook to a new file
    excel_path_with_images = './ozon_relate/ozon_with_images_.xlsm'
    wb.save(excel_path_with_images)

if __name__ == '__main__':
    asyncio.run(main())
