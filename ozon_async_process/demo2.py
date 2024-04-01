# encoding='utf-8
# @Time: 2023-12-30
# @File: %
#!/usr/bin/env

from curl_cffi import requests
import requests as req
from lxml import etree
import logging
from parseData import *
import concurrent.futures
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO
from read_file import ReadFile
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import ProcessPoolExecutor
import multiprocessing

from cookie_pool import GenCookie


class Ozon_Spider:
    def __init__(self, url):
        self.session = requests.Session()
        self.url = url
        self.headers = {
            # ... (same as before)
        }
        self.session.headers.update(self.headers)

    def searchResultsV2(self, cookies, ua, headers):
        try:
            response = self.session.get(self.url, cookies=cookies, headers=headers, impersonate=ua, timeout=30)
            html = etree.HTML(response.text)
            data = html.xpath('//div[@class="jn6 j6n"]//img/@src')
            if not data:
                data = html.xpath('//div[@class="rj"]/img/@src')
                if not data:
                    data = html.xpath('//link[@rel="preload"]/@href')
            return data[0] if data else None
        except Exception as e:
            logging.error(f'Error getting image URL: {e}')
            return None

def download_image(url, headers, index, ws_data):
    try:
        if url:
            img_url = url.replace('wc50', 'wc1000')
            logging.info(f'Downloading image {index-1}: {img_url}')
            with requests.Session() as session:
                response = session.get(img_url, headers=headers, timeout=30)
                img = Image.open(BytesIO(response.content))
                img = img.resize((150, 150), Image.LANCZOS)
                img_byte_arr = BytesIO()
                img.save(img_byte_arr, format='PNG')
                ws_data[index] = (BytesIO(img_byte_arr.getvalue()), img_url)
    except Exception as e:
        logging.error(f'Error downloading image {index-1}: {e}')

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')
    gencookies = GenCookie()
    cookies, ua, headers = gencookies.gen_cookie()
    file_path = r"./ozon_imgdown/data.xlsx"
    # save_path = r"./ozon_imgdown/img.xlsx"
    read_file = ReadFile(file_path)
    df = read_file.df
    wb = read_file.wb
    ws = read_file.ws
    max_workers = os.cpu_count() * 5
    
    manager = multiprocessing.Manager()
    ws_data = manager.dict()
    
    with ThreadPoolExecutor() as executor:
        futures = []
        for index, row in enumerate(df.itertuples(), start=2):
            url = getattr(row, '商品链接')
            ozon = Ozon_Spider(url)
            img_url = ozon.searchResultsV2(cookies, ua, headers)
            futures.append(executor.submit(download_image, img_url, headers, index, ws_data))

        for future in concurrent.futures.as_completed(futures):
            future.result()

    for index, (img_data, img_url) in ws_data.items():
        openpyxl_img = OpenpyxlImage(img_data)
        cell = 'A' + str(index)
        ws.add_image(openpyxl_img, cell)
        ws.cell(row=index, column=2).value = img_url

    # Set column width and row height
    ws.column_dimensions[get_column_letter(1)].width = 20
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 150 * 0.75

    # Save the workbook to a new file
    excel_path_with_images = './ozon_imgdown/ozon_with_images.xlsm'
    wb.save(excel_path_with_images)
