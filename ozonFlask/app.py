# encoding='utf-8
# @Time: 2024-01-06
# @File: %
#!/usr/bin/env
from icecream import ic
import os
from flask import Flask, render_template, request, session, redirect, url_for
from pymongo import MongoClient
from flask import Flask, send_file
from openpyxl import Workbook
import os

app = Flask(__name__)
# 连接到MongoDB
client = MongoClient("mongodb://127.0.0.1:27017/")
db = client['ozon']
collection = db['ozon']
app.secret_key = '123456'  # 设置Flask会话的秘密密钥

@app.route('/')
def index():
    # 获取页码和每页显示条目数
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))

    # MongoDB的分页查询
    data = list(collection.find().skip((page - 1) * per_page).limit(per_page))
    ic(data)
    # 总文档数，用于计算总页数
    session['search_query'] = {}
    total_count =  collection.count_documents({})
    total_pages = (total_count + per_page - 1) // per_page
    session['search_query'] = {}

    # 渲染页面并传递数据
    return render_template('index.html', data=data, page=page, total_pages=total_pages, per_page=per_page)

@app.route('/result')
def result():
    # 默认的分页参数
    per_page = int(request.args.get('per_page', 50))
    page = int(request.args.get('page', 1))
    # ic(page)
    # 从会话中获取保存的查询条件
    query = session.get('search_query', {})

    # 根据查询条件获取数据
    total_count = collection.count_documents(query)
    total_pages = (total_count + per_page - 1) // per_page

    result = list(collection.find(query).skip((page - 1) * per_page).limit(per_page))

    return render_template('result.html', result=result, page=page, total_pages=total_pages, per_page=per_page)



@app.route('/search', methods=['GET', 'POST'])
def search():
    # 默认的分页参数
    per_page = 50
    page = int(request.args.get('page', 1))

    if request.method == 'POST':
        # 从表单获取数据
        sku_ids_str = request.form.get('sku_ids')
        search_text = request.form.get('search_text')
        sku_id1 = request.form.get('skuId1')
        search_text1 = request.form.get('search_text1')

        # 拆分并创建查询条件
        query = {}
        if sku_ids_str:
            sku_ids = sku_ids_str.splitlines()
            query["skuId"] = {"$in": sku_ids}
        if search_text:
            search_texts = search_text.splitlines()
            query["SearchText"] = {"$in": search_texts}
        if sku_id1 and search_text1:
            sku_id1_list = sku_id1.splitlines()
            search_text1_list = search_text1.splitlines()
            query = {"skuId": {"$in": sku_id1_list}, "SearchText": {"$in": search_text1_list}}

        # 保存查询条件到会话
        session['search_query'] = query

        # 重定向到新的路由
        return redirect(url_for('result'))
    
    return render_template('search.html')

    
@app.route('/download')
def download_file():
    # 查询数据
    documents = session.get('search_query')
    # ic(documents)
    documents = list(collection.find(documents))
 
    
    # ic(documents)
    # 创建一个Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 假设我们要写入的数据包含以下字段
    fields = ["ImagUrl","skuId","title","url","price","rating","review","MaxAddToCart","SearchText","SearchDate","pageRank","Advert","page","Rank"]  # 替换为实际字段名

    # 写入标题行
    for i, field in enumerate(fields, 1):
        ws.cell(row=1, column=i, value=field)

    # 写入数据行
    for row_index, document in enumerate(documents, 2):
        for col_index, field in enumerate(fields, 1):
            ws.cell(row=row_index, column=col_index, value=document.get(field, ''))

    # 设置文件名和保存路径
    filename = 'my_excel_file.xlsm'
    filepath = os.path.join('C:\\apps\\', filename)

    # 保存工作簿
    wb.save(filepath)

    # 提供文件下载
    return send_file(filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True,port=5002)


