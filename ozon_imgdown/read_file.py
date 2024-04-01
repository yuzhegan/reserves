# encoding='utf-8

# @Time: 2024-03-30
# @File: %
#!/usr/bin/env
import asyncio
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class AsyncReadFile:
    def __init__(self, file_path):
        self.file_path = file_path

    async def read_file(self):
        loop = asyncio.get_event_loop()
        self.df = await loop.run_in_executor(None, pd.read_excel, self.file_path)
        self.wb = await loop.run_in_executor(None, load_workbook, self.file_path, read_only=True)
        self.ws = self.wb.active
        for r in dataframe_to_rows(self.df, index=False, header=True):
            self.ws.append(r)
        await self.add_column()
        return self.df, self.wb, self.ws

    async def add_column(self):
        self.ws.insert_cols(1)
        self.ws.cell(row=1, column=1).value = '图片连接'
        self.ws.insert_cols(1)
        self.ws.cell(row=1, column=1).value = '图片'
